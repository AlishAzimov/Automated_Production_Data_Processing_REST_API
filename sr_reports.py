import os
import requests
import pandas as pd
import numpy as np
import urllib3
from tqdm import tqdm
from datetime import date,timedelta
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
import argparse



# Ссылка на API
BASE_URL = "######"
TOKEN = "#####"
# Исключает ошибку с не доверием urllib3 к API
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Глобальные переменные 
country_id = 3              # Регион "Средняя Азия"

# Подключение к API
session = requests.Session()
session.headers.update({
    "Authorization": TOKEN,
    "Accept": "application/json"
})


# Функция для получение справочной ифномрации
def get_reference(query, id_list=None, is_dvd=False): 
    
    result = []

    # Если список ID не передан → запрос всей таблиц
    if id_list is None:
        try:
            response = session.get(
                f"{BASE_URL}/{query}",  # без ID
                verify=False
            )
            response.raise_for_status()
            data = response.json()
        except Exception as e:
            print(f"Ошибка при запросе всей таблицы {query}: {e}")
            return []
        
        for d in tqdm(data, desc=f"📥 Загрузка всей таблицы {query}", unit="запись"):
            result.append(d)

        return result

    # Если список ID есть → поштучно
    for id in tqdm(id_list, desc=f"📥 Загрузка справок по {query}", unit="запись"):
        # проверка на NaN
        if pd.isna(id):
            continue
        

        try:
            if not isinstance(id, int):
                raise TypeError(f"type id должно быть int, а не {type(id)}")
            response = session.get(
            f"{BASE_URL}/{query}/{id}",
            verify=False
            )
            response.raise_for_status()  # выбросит ошибку, если статус не 200
            data = response.json()
        except Exception as e:
            print(f"Ошибка при запросе {id}: {e}")
            continue

        if not data:
            continue
 
        if is_dvd:
        # API возвращает список словарей
            if isinstance(data, list):
                for d in data:
                    d["well_id"] = id
                result.extend(data)
            # API возвращает одиночный словарь
            elif isinstance(data, dict):
                data["well_id"] = id
                result.append(data)
        else:
            result.append(data)   


    return result



# Функция для получение суточных рапортов 
def get_timedata():
    try:
        timedata =[]
        page=1 
        max_pages=1000
        yesterday = date.today() - timedelta(days=1)
        date_time = date_to if pd.to_datetime(date_to) < pd.to_datetime(yesterday) else yesterday 
        
        with tqdm(desc="📥 Загрузка рапортов по работам", unit=" шт.") as pbar:
            while True:  
                if page>=max_pages:
                    if timedata:
                        print(f"Выполнен лимит: {max_pages} страниц, "
                            f"получено {len(timedata)} суточных рапортов, "
                            f"дата последнего рапорта {timedata[0]['date']}")
                    else:
                        print(f"Выполнен лимит: {max_pages} страниц, данных нет.")
                    break

                response = session.get(
                    f"{BASE_URL}/time-data",
                    params={
                        "page":page,
                        "date_from":date_from,    # Дата начало рапорта
                        "country_id": country_id  # Регион "Средняя Азия"
                    },
                    verify=False,
                    stream=True
                )

                before = len(timedata)

                if response.status_code != 200:
                    print(f"Ошибка запроса: {response.status_code} - {response.text}")
                    return None

                data = response.json() 
                if not data:
                    break



                if pd.to_datetime(data[-1]['date']) < pd.to_datetime(date_time):
                        timedata.extend(data)
                else:
                    for item in data:
                        if pd.to_datetime(item['date']) <= pd.to_datetime(date_time):
                            timedata.append(item)
                        else:
                            break
                    added = len(timedata) - before
                    if added > 0:
                        pbar.update(added)
                    break


                added = len(timedata) - before
                if added > 0:
                    pbar.update(added)          

                page+=1

            
        # перевоДим в DataFrame

        timedata=pd.DataFrame(timedata)
        timedata['date']=pd.to_datetime(timedata['date'])
        timedata['start']=pd.to_datetime(timedata['start'])
        timedata['end']=pd.to_datetime(timedata['end'])
        timedata['passage_trunk'] = pd.to_numeric(timedata['passage_trunk'], errors='coerce')
        timedata['passage'] = pd.to_numeric(timedata['passage'], errors='coerce')
        timedata['npt_type']=timedata['npt_type'].map({1: "собственное", 2: "прочие"})
        timedata=timedata.rename(columns={
        'id': 'timeDataID',
        'date': 'reportDate',
        'passage_trunk':'забой_конце_рапорта_метр',
        'start':'дата_начала_работ',
        'end':'дата_окончания_работ',
        'type_id':'timeDataType_id',
        'passage':'забой_конце_работы_метр',
        'seconds':'длительность_секунд',
        'npt_type':'тип_НПВ',
        'comment':'описание_работ'
        })
        

        timeDataTypes=[]
        timeDataType_id=timedata['timeDataType_id'].dropna().unique().astype(int).tolist()
        timeDataTypes=get_reference('time-data-types',timeDataType_id)
        timeDataTypes=pd.DataFrame(timeDataTypes)
        timeDataTypes=timeDataTypes.rename(columns={
        'id':'timeDataType_id',
        'name': 'тип_работ',
        'isNPT':'НПВ',
        'isNC':'некалендарное_время'})
        timeDataTypes=timeDataTypes.drop(columns=['parent'])

        departments=[]
        departments=get_reference('departments')
        departments=pd.DataFrame(departments)
        departments=departments.rename(columns={
        'id':'guilty',
        'name': 'НПВ_ответственный'})
        departments=departments.drop(columns=['comment'])
        
        timedata = timedata.merge(timeDataTypes, on="timeDataType_id", how="left")
        timedata = timedata.merge(departments, on="guilty", how="left")


        timedata=timedata.drop(columns=['customer_id','drilling_rig_id','well_construction_element_is_absent',
                                        'timeDataType_id','guilty','timeDataID'])


       
        return timedata
    
    except Exception as e:
        print(f"Ошибка при получении данных: {e}")
        return None
    





# Функция для получение и сбора справочной инфомраций 
def get_wells(well_id):
    
       
    # Получаем инфомацию о скважинах
    weels = []
    weels=get_reference('wells',well_id)
    weels=pd.DataFrame(weels)
    weels['date_start']=pd.to_datetime(weels['date_start'])
    weels['date_end']=pd.to_datetime(weels['date_end'])
    weels=weels.rename(columns={
    'id': 'well_id',
    'date_start': 'wellStartDate',
    'date_end': 'wellEndDate',
    })
    wellConstruction = weels.explode('wellConstructionElements', ignore_index=True)
    weels = weels[["well_id","name","cluster_id","is_functioning","purpose_id","customer_id", 
                   "general_contractor_id","wellStartDate", "wellEndDate","days_planned","drilling_rig_id"]]

     # Получаем инфомацию о кусте скважины
    clusters = []
    cluster_id=weels['cluster_id'].dropna().unique().astype(int).tolist()
    clusters=get_reference('clusters',cluster_id)
    clusters=pd.DataFrame(clusters)
    clusters=clusters.rename(columns={
    'id':'cluster_id',
    'name': 'куст'})

    # Получаем инфомацию о типе скважины
    purposes = []
    purpose_id=weels['purpose_id'].dropna().unique().astype(int).tolist()
    purposes=get_reference('purposes',purpose_id)
    purposes=pd.DataFrame(purposes)
    purposes=purposes.rename(columns={
    'id':'purpose_id',
    'name': 'тип_скважины'})

    # Получаем инфомацию о заказчике
    customers = []
    customer_id=weels['customer_id'].dropna().unique().astype(int).tolist()
    customers=get_reference('customers',customer_id)
    customers=pd.DataFrame(customers)
    customers=customers.rename(columns={
    'id':'customer_id',
    'name': 'заказчик'})
    customers = customers[['customer_id','заказчик']]

    # Получаем инфомацию о ген_подрятчике
    contractors = []
    general_contractor_id=weels['general_contractor_id'].dropna().unique().astype(int).tolist()
    contractors=get_reference('contractors',general_contractor_id)
    contractors=pd.DataFrame(contractors)
    contractors=contractors.rename(columns={
    'id':'general_contractor_id',
    'name': 'ген_подрятчик'})
    contractors = contractors[['general_contractor_id','ген_подрятчик']]

    # Получаем инфомацию о буровой установке
    drilling_rigs = []
    drilling_rig_id=weels['drilling_rig_id'].dropna().unique().astype(int).tolist()
    drilling_rigs=get_reference('drilling-rigs',drilling_rig_id)
    drilling_rigs=pd.DataFrame(drilling_rigs)
    drilling_rigs=drilling_rigs.rename(columns={
    'id':'drilling_rig_id',
    'type': 'тип_БУ',
    'code': 'код_БУ',
    'number': 'номмер_БУ',
    'year': 'год_выпуска_БУ'
    })
    drilling_rigs = drilling_rigs.drop(columns=["comment","history"])

    # Получаем инфомацию о месторождений
    fields = []
    field_id=clusters['field_id'].dropna().unique().astype(int).tolist()
    fields=get_reference('fields',field_id)
    fields=pd.DataFrame(fields)
    fields=fields.rename(columns={
    'id':'field_id',
    'name': 'месторождение'})
     
    # Получаем инфомацию о регионе
    regions = []
    region_id=fields['region_id'].dropna().unique().astype(int).tolist()
    regions=get_reference('regions',region_id)
    regions=pd.DataFrame(regions)
    regions=regions.rename(columns={
    'id':'region_id',
    'name': 'регион'})
    regions = regions.drop(columns=["country_id"])

    # Элемент контсрукции
    elements = pd.json_normalize(wellConstruction['wellConstructionElements'])
    element_id=elements['element_id'].dropna().unique().astype(int).tolist()
    construction_elements = get_reference('construction-elements',element_id)
    construction_elements=pd.DataFrame(construction_elements)
    construction_elements=construction_elements.rename(columns={
    'id':'element_id',
    'name':'колонна'})
    construction_elements=construction_elements[['element_id','колонна']]
    elements=elements.rename(columns={
    'id':'well_construction_element_id',
    'head_plan':'голова_ОК_план',
    'shoe_plan':'башмак_ОК_план',
    'head_fact':'голова_ОК_факт',
    'shoe_fact':'башмак_ОК_факт',
    'start_date':'дата_начало_этап_ОК',
    'end_date':'дата_окончания_этап_ОК'})
    elements=elements.merge(construction_elements,on="element_id", how="left")
    elements=elements.drop(columns=["element_id","sort","date"])

    # Собираем все таблицы в одну
    result = weels.merge(purposes, on="purpose_id", how="left")
    result = result.merge(customers, on="customer_id", how="left")
    result = result.merge(clusters, on="cluster_id", how="left")
    result = result.merge(contractors, on="general_contractor_id", how="left")
    result = result.merge(drilling_rigs, on="drilling_rig_id", how="left")
    result = result.merge(fields, on="field_id", how="left")
    result = result.merge(regions, on="region_id", how="left")



    result = result.drop(columns=["purpose_id","customer_id","cluster_id","general_contractor_id",
                                  "drilling_rig_id","field_id","region_id"])

    return result, elements




# Функция для получение план бурение
def get_dvd(well_id):
    dvd = []
    dvd=get_reference('dvd',well_id,is_dvd=True)
    dvd=pd.DataFrame(dvd)
    dvd=dvd.rename(columns={
    'id': 'dvd_id',
    'duration': 'план_длительность_секунд',
    'passage': 'план_забой',
    'sort': 'dvd_sort',
    'description':'план_описание_работ',
    'time_data_type_id':'timeDataType_id'
    })

    timeDataTypes=[]
    timeDataType_id=dvd['timeDataType_id'].dropna().unique().astype(int).tolist()
    timeDataTypes=get_reference('time-data-types',timeDataType_id)
    timeDataTypes=pd.DataFrame(timeDataTypes)
    timeDataTypes=timeDataTypes.rename(columns={
    'id':'timeDataType_id',
    'name': 'план_тип_работ'})
    timeDataTypes=timeDataTypes.drop(columns=['isNPT','parent','isNC'])

    dvd = dvd.merge(timeDataTypes, on="timeDataType_id", how="left")
    dvd=dvd.drop(columns=['timeDataType_id'])

    return dvd
    

#добвление проходки
def add_passage (df,isdvd=False):
    WELL = 'well_id'
    SDATE='wellStartDate'
    WCEI='well_construction_element_id'
    DVDS='dvd_sort'
    PLANS='план_длительность_секунд'
    DEPTH = 'забой_конце_работы_метр' if not isdvd else 'план_забой'
    DATE = 'дата_начала_работ'
    PENET = 'проходка_конце_работы_метр'


    key = [WELL, DATE] if not isdvd else [WELL, WCEI,DVDS]
    key2 = WELL if not isdvd else [WELL,WCEI]

    # сохраняем исходный порядок строк
    df["_orig_order"] = range(len(df))
    
   
    # сортировка по скважине и дате
    df = df.sort_values(key, kind='mergesort')

    if isdvd:
        t0 = df.groupby([WELL,WCEI])[SDATE].transform('min')
        cum_sec = df.groupby([WELL,WCEI])[PLANS].cumsum()
        df[DATE] = t0 + pd.to_timedelta(cum_sec, unit='s')
    
    # предыдущее значение забоя по каждой скважине
    prev = df.groupby(key2)[DEPTH].shift(1)
    delta = (df[DEPTH] - prev).fillna(0)
    df[PENET] = delta.clip(lower=0) if isdvd else delta

    # возвращаем обратно к исходному порядку
    df = df.sort_values("_orig_order").drop(columns="_orig_order")

    return df

# создание итоговой сводной таблицы для план
def totals_group_plan(dataf,grpColumn):
    df=dataf[dataf['дата_начала_работ']>=pd.to_datetime(date_from)].copy()

    RDate='дата_начала_работ' 

    df['Год']=df[RDate].dt.year
    df['Месяц']=df[RDate].dt.month
    df['Квартал'] = df[RDate].dt.quarter
    df = df[df['заказчик'] != 'СП ООО "Gissarneftegaz"'].copy()

    cond1 = df['план_тип_работ'].eq('Сплошное бурение')
    cond2 = df['план_описание_работ'].str.startswith('Бурение в интервале', na=False)
    mask = cond1 | cond2
    df['длительность_сек_сплошное_бурение'] = df['план_длительность_секунд'].where(mask, 0)  

    keys = ['Год','Квартал']
    keys.extend(grpColumn)
    
    res = (
        df.groupby(keys)
        .agg(
            уникальные_БУ=('номмер_БУ', 'nunique'),
            уникальные_скваж=('well_id', 'nunique'),
            проходка_м=('проходка_конце_работы_метр', 'sum'),
            календарное_время_cут=('план_длительность_секунд', lambda s: s.sum() / 86400),
            время_cплошное_бурение_час=('длительность_сек_сплошное_бурение', lambda s: s.sum() / 3600),
            
        )
        .assign(коммерческая_скорость_м_сут=lambda x: x['проходка_м'] / x['календарное_время_cут'],
                механическая_скорость_м_час=lambda x: x['проходка_м'] / x['время_cплошное_бурение_час'])
        .reset_index()
    )
    res=res.replace([np.inf, -np.inf], np.nan)
    

    return res

# создание итоговой сводной таблицы для факт
def totals_group_fact(dataf,grpColumn):
    df=dataf.copy()

    RDate='reportDate' 

    df['Год']=df[RDate].dt.year
    df['Месяц']=df[RDate].dt.month
    df['Квартал'] = df[RDate].dt.quarter
    df = df[df['заказчик'] != 'СП ООО "Gissarneftegaz"'].copy()
    mask = df['тип_работ'].eq('Сплошное бурение')
    df['длительность_сек_сплошное_бурение'] = df['длительность_секунд'].where(mask, 0)   
    df['длительность_НПВ'] = df['длительность_секунд'].where(df['НПВ'], 0)
    
    keys = ['Год','Квартал']
    keys.extend(grpColumn)
    
    res = (
        df.groupby(keys)
        .agg(
            уникальные_БУ=('номмер_БУ', 'nunique'),
            уникальные_скваж=('well_id', 'nunique'),
            проходка_м=('проходка_конце_работы_метр', 'sum'),
            календарное_время_cут=('длительность_секунд', lambda s: s.sum() / 86400),
            время_cплошное_бурение_час=('длительность_сек_сплошное_бурение', lambda s: s.sum() / 3600),
            время_НПВ_час=('длительность_НПВ', lambda s: s.sum() / 3600),
            
        )
        .assign(коммерческая_скорость_м_сут=lambda x: x['проходка_м'] / x['календарное_время_cут'],
                механическая_скорость_м_час=lambda x: x['проходка_м'] / x['время_cплошное_бурение_час'])
        .reset_index()
    )
    res=res.replace([np.inf, -np.inf], np.nan)

    return res

# создание сводной таблицы по НПВ
def pivot_table_npv(dataf):
    df=dataf.copy()

    RDate='reportDate' 

    df['Год']=df[RDate].dt.year
    df['Месяц']=df[RDate].dt.month
    df['Квартал'] = df[RDate].dt.quarter

    df_npv = df[df['НПВ'] == True].copy()
    df_npv['Часы_НПВ'] = df_npv['длительность_секунд'] / 3600.0


    # сводная таблица сумма часов НПВ по регионам
    pivot_sum_by_region = pd.pivot_table(
        df_npv,
        values='Часы_НПВ',
        index=['Год', 'Квартал', 'регион'],
        columns=['НПВ_ответственный', 'тип_НПВ'],
        aggfunc='sum',
        fill_value=0,
        margins=True,
        margins_name='Итого'   
    )

    pivot_sum_by_region=pivot_sum_by_region.round(1)
    pivot_sum_by_region.columns.names = ["Ответственный за НПВ", "Тип НПВ"]
    pivot_sum_by_region = pd.concat({'Итоговое распределение часов НПВ по регионам': pivot_sum_by_region}, axis=1)

    # сводная таблица сумма часов НПВ по заказчикам
    pivot_sum_by_customer = pd.pivot_table(
        df_npv,
        values='Часы_НПВ',
        index=['Год', 'Квартал', 'заказчик'],
        columns=['НПВ_ответственный', 'тип_НПВ'],
        aggfunc='sum',
        fill_value=0,
        margins=True,
        margins_name='Итого'    
    )

    pivot_sum_by_customer=pivot_sum_by_customer.round(1)
    pivot_sum_by_customer.columns.names = ["Ответственный за НПВ", "Тип НПВ"]
    pivot_sum_by_customer = pd.concat({'Итоговое распределение часов НПВ по заказчикам': pivot_sum_by_customer}, axis=1)

    # Сводная таблица сумма часов НПВ по заказчикам с детализацией по ответственным ДМТО
    df_dmto=df_npv[df_npv['НПВ_ответственный']=='ДМТО'].copy()
    pivot_dmto_sum_by_customer = pd.pivot_table(
        df_dmto,
        values='Часы_НПВ',
        index=['Год', 'Квартал', 'заказчик'],
        columns=['тип_работ'],
        aggfunc='sum',
        fill_value=0,
        margins=True,
        margins_name='Итого'    
    )

    pivot_dmto_sum_by_customer=pivot_dmto_sum_by_customer.round(1)
    pivot_dmto_sum_by_customer.columns.names = ["Тип работ НПВ"]
    pivot_dmto_sum_by_customer = pd.concat({'Распределение часов НПВ по заказчикам с детализацией по ответственным ДМТО': pivot_dmto_sum_by_customer}, axis=1)

    # Сводная таблица сумма часов НПВ по заказчикам с детализацией по ответственным ДМТО
    df_fin=df_npv[df_npv['НПВ_ответственный']=='Финансовый Департамент'].copy()
    pivot_fin_sum_by_customer = pd.pivot_table(
        df_fin,
        values='Часы_НПВ',
        index=['Год', 'Квартал', 'заказчик'],
        columns=['тип_работ'],
        aggfunc='sum',
        fill_value=0,
        margins=True,
        margins_name='Итого'   
        )

    pivot_fin_sum_by_customer=pivot_fin_sum_by_customer.round(1)
    pivot_fin_sum_by_customer.columns.names = ["Тип работ НПВ"]
    pivot_fin_sum_by_customer = pd.concat({'Распределение часов НПВ по заказчикам с детализацией по ответственным Финансовым Департаментом': pivot_fin_sum_by_customer}, axis=1)


    return pivot_sum_by_region, pivot_sum_by_customer, pivot_dmto_sum_by_customer, pivot_fin_sum_by_customer


# ─────────────────────────────────────────────────────────────────────────────
# Автооформление таблицы (одной) на листе
def _autoformat_sheet(ws, df: pd.DataFrame):
    # БЕЗ автофильтра и БЕЗ заморозки
    # ws.auto_filter.ref = ws.dimensions
    # ws.freeze_panes = "A2"

    # Шапка
    header_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    # Авто-ширина
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 60)


# ─────────────────────────────────────────────────────────────────────────────
# Универсальная запись нескольких блоков (DataFrame'ов) на один лист
def _write_blocks_to_sheet(writer, sheet_name: str, blocks, gap_rows: int = 3, bold_titles: bool = True):
    """
    blocks: список (title: str, df: pd.DataFrame)
    """
    # создать пустой лист и очистить
    pd.DataFrame({"_": []}).to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]
    ws.delete_rows(1, ws.max_row)

    cur = 0
    title_font = Font(bold=True) if bold_titles else Font(bold=False)

    for title, df in blocks:
        if df is None or (hasattr(df, "empty") and df.empty):
            continue

        # Заголовок блока
        ws.cell(row=cur + 1, column=1, value=title).font = title_font

        # Таблица начинается СЛЕДУЮЩЕЙ строкой после заголовка
        startrow = cur + 2
        # index=True оставляем как есть (pandas сам решит по df)
        df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=True)

        # После записи берём фактический последний ряд и добавляем зазор
        cur = ws.max_row + gap_rows

    # БЕЗ автофильтра и БЕЗ заморозки
    # ws.auto_filter.ref = ...
    # ws.freeze_panes = ...

    # Авто-ширина по всему листу
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 60)

    return ws


# ─────────────────────────────────────────────────────────────────────────────
def save_report(
    Reports: pd.DataFrame,
    Dvd: pd.DataFrame,
    date_from: str,
    # Пивоты по НПВ:
    PivotSumByRegion: pd.DataFrame = None,
    PivotSumBycustomer: pd.DataFrame = None,
    PivotDMTOSumByCustomer: pd.DataFrame = None,
    PivotFinSumByCustomer: pd.DataFrame = None,
    # Итоги план/факт:
    TotalFactCustomer: pd.DataFrame = None,
    TotalFactRegion: pd.DataFrame = None,
    TotalPlanCustomer: pd.DataFrame = None,
    TotalPlanRegion: pd.DataFrame = None,
    outdir: str = "data",
    filename_tpl: str = "ReportsDateFrom{date_from}.xlsx",
    flatten_pivots: bool = False,   # если True — сбрасываем индексы у сводных для «плоской» выгрузки
):
    """
    Сохраняет отчёт в Excel с листами и порядком:
      1. «Свод по НПВ» — PivotSumByRegion, PivotSumBycustomer, PivotDMTOSumByCustomer, PivotFinSumByCustomer
      2. «Итоги план/факт» — TotalFactCustomer, TotalFactRegion, TotalPlanCustomer, TotalPlanRegion
      3. «РаспределениеСуточныхРапортов» — Reports
      4. «РаспределениеПлан» — Dvd
    """
    os.makedirs(outdir, exist_ok=True)
    path = os.path.join(outdir, filename_tpl.format(date_from=date_from))

    # При необходимости «приплюснём» пивоты (многомерные индексы) в плоские таблицы
    def _maybe_flatten(df):
        if df is None:
            return df
        if not flatten_pivots:
            return df
        out = df.copy()
        if isinstance(out.index, pd.MultiIndex):
            out = out.reset_index()
        if isinstance(out.columns, pd.MultiIndex):
            out.columns = [' | '.join([str(x) for x in col]).strip() for col in out.columns.to_flat_index()]
        return out

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        # 1) Свод по НПВ — с зазором 3 строки
        _write_blocks_to_sheet(
            writer,
            sheet_name="Свод по НПВ",
            blocks=[
                ("Итоговое распределение часов НПВ по регионам", PivotSumByRegion),
                ("Итоговое распределение часов НПВ по заказчикам", PivotSumBycustomer),
                ("Часы НПВ: вклад ответственных ДМТО по заказчикам", PivotDMTOSumByCustomer),
                ("Финансовые НПВ по заказчикам", PivotFinSumByCustomer),
            ],
            gap_rows=3,
            bold_titles=True,
        )

        # 2) Итоги план/факт — тоже с зазором (для единообразия)
        _write_blocks_to_sheet(
            writer,
            sheet_name="Итоги план—факт",  # или прогони через _safe_sheet_name
            blocks=[
                ("ФАКТ: суммарно по заказчикам", TotalFactCustomer),
                ("ФАКТ: суммарно по регионам", TotalFactRegion),
                ("ПЛАН: суммарно по заказчикам", TotalPlanCustomer),
                ("ПЛАН: суммарно по регионам", TotalPlanRegion),
            ],
            gap_rows=3,
            bold_titles=True,
        )

        # 3) и 4) — просто выгружаем и автоформатим (БЕЗ фильтров/заморозки)
        Reports.to_excel(writer, sheet_name="РаспределениеСуточныхРапортов", index=False)
        _autoformat_sheet(writer.sheets["РаспределениеСуточныхРапортов"], Reports)

        Dvd.to_excel(writer, sheet_name="РаспределениеПлан", index=False)
        _autoformat_sheet(writer.sheets["РаспределениеПлан"], Dvd)

    return path



if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Выгрузка отчетов из **** в Excel\n"
                    "Пример использования:\n"
                    "  python sr_reports.py 2025-08-01 2025-08-20\n"
                    "  python sr_reports.py 2025-08-01",
        formatter_class=argparse.RawTextHelpFormatter
    )

    parser.add_argument(
        "date_from",
        help="Дата начала отчета (включительно, формат YYYY-MM-DD, пример: 2025-08-01)"
    )
    parser.add_argument(
        "date_to",
        nargs="?",
        default=None,
        help="Дата конца отчета (включительно, формат YYYY-MM-DD, пример: 2025-08-20). "
             "Если не указана, берется сегодняшняя дата."
    )

    args = parser.parse_args()

    # Глобальные переменные
    date_from = args.date_from
    date_to = args.date_to if args.date_to else str(date.today())

    print(f"📅 Период: {date_from} → {date_to}")
    print("──────────────")
    # Запуск функций
    Reports = get_timedata()
    well_id = Reports['well_id'].dropna().unique().astype(int).tolist()
    Wells, Elements=get_wells(well_id)
    Dvd = get_dvd(well_id)

    print("🔄 Сборка всех данных...")
   
    Reports = Reports.merge(Wells, on="well_id", how="left")
    Dvd = Dvd.merge(Wells, on="well_id", how="left")
    Reports=Reports.merge(Elements, on="well_construction_element_id", how="left")
    Dvd=Dvd.merge(Elements, on="well_construction_element_id", how="left")
    Reports = Reports.sort_values(["report_id", "дата_начала_работ"],ascending=False)
    Dvd = Dvd.sort_values(["dvd_id", "dvd_sort"],ascending=False)
    
    print("✅ Данные успешно собраны!")
    print("──────────────")
    print("📊 Формирование отчётов...")
    
    # Создание столбца проходка
    Reports=add_passage(Reports)
    Dvd=add_passage(Dvd, isdvd=True)
    
    Reports.to_csv("data/Reports.csv", index=False, encoding="utf-8-sig")
    Dvd.to_csv("data/Dvd.csv", index=False, encoding="utf-8-sig")

    # Создание итогов план/проект
    TotalFactCustomer=totals_group_fact(Reports,['заказчик'])
    TotalFactRegion=totals_group_fact(Reports,['регион'])
    TotalPlanCustomer=totals_group_plan(Dvd,['заказчик'])
    TotalPlanRegion=totals_group_plan(Dvd,['регион'])

    # Создание сводных по НПВ
    PivotSumByRegion, PivotSumBycustomer, PivotDMTOSumByCustomer, PivotFinSumByCustomer=pivot_table_npv(Reports)


    print("💾 Сохранение отчета...")
    path = save_report(
    Reports,
    Dvd,
    date_from,
    PivotSumByRegion=PivotSumByRegion,
    PivotSumBycustomer=PivotSumBycustomer,
    PivotDMTOSumByCustomer=PivotDMTOSumByCustomer,
    PivotFinSumByCustomer=PivotFinSumByCustomer,
    TotalFactCustomer=TotalFactCustomer,
    TotalFactRegion=TotalFactRegion,
    TotalPlanCustomer=TotalPlanCustomer,
    TotalPlanRegion=TotalPlanRegion,
    flatten_pivots=False  # поставь True, если надо выгружать «плоские» сводные
    )

    print(f"✅ Отчет успешно сохранен: {path}")
